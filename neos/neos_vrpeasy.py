from lrp_easy import createLRP
from solver_cvrp import dataCvrp
import os
import openpyxl
from dataparse import create_data
from datetime import datetime
import logging
import sys

log_dir = "log_files/mip_nn"
os.makedirs(log_dir, exist_ok=True)

# Directory containing the prodhon dataset
directory_path = "/Users/waquarkaleem/NEOS-LRP-Codes-2/prodhon_dataset"

# Directory to store decision informed instances 
DIL_instances = "/Users/waquarkaleem/NEOS-LRP-Codes-2/neos/dil_instances"

# Write the results in the excel
existing_excel_file="/Users/waquarkaleem/NEOS-LRP-Codes-2/results/DFL_LRP.xlsx" 
workbook = openpyxl.load_workbook(existing_excel_file)
sheet_name = "iter_1_systematic_bs1_50frompool"

if sheet_name not in workbook.sheetnames:
    workbook.create_sheet(sheet_name)
worksheet = workbook[sheet_name]

def write_to_txt_cvrplib_format(depot_id, depot_customers, depot_coords, customer_demands, filename, vehicle_capacity):
    with open(filename, 'w') as file:
        file.write(f"NAME : {os.path.basename(filename)}\n")
        file.write("COMMENT : decision informed instance\n")
        file.write("TYPE : CVRP\n")
        file.write(f"DIMENSION : {len(depot_customers) + 1}\n")  # +1 for the depot
        file.write("EDGE_WEIGHT_TYPE : EUC_2D\n")
        file.write(f"CAPACITY : {vehicle_capacity}\n")
        file.write("NODE_COORD_SECTION\n")
        
        file.write(f"1 {depot_coords[0][0]} {depot_coords[0][1]}\n")
        
        for i, coords in enumerate(depot_customers, start=2):
            file.write(f"{i} {coords[0]} {coords[1]}\n")

        file.write("DEMAND_SECTION\n")
        
        for i, demand in enumerate(customer_demands, start=1):
            file.write(f"{i} {demand}\n")

        file.write("DEPOT_SECTION\n")
        file.write("1\n")  
        file.write("-1\n")  
        file.write("EOF\n")

for filename in os.listdir(directory_path):
    if filename.endswith(".dat"):  # Adjust the file extension as needed
        file_path = os.path.join(directory_path, filename)
        print("Working on :",file_path)
        
        # Call your function to get data
        # Create the log file name based on the input file's name
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  
        log_filename = f"{os.path.splitext(filename)[0]}_{current_time}.log"

        logging.basicConfig(
            level=logging.INFO,  # Set the logging level (INFO, WARNING, ERROR, etc.)
            format='%(asctime)s - %(levelname)s - %(message)s',  # Define the log message format
            handlers=[
                logging.FileHandler(os.path.join(log_dir, log_filename)),  # Log to a file
                logging.StreamHandler(sys.stdout)   # Log to stdout
            ]
        )
        logging.info(f'\n\n Working on file :{file_path}')
        ans = create_data(file_path)  # Assuming create_data is a function that processes the data
        lrp_st=datetime.now()
        lrp_solver = createLRP(ans)  # Assuming createFlp is a function that creates the solver
        lrp_result=lrp_solver.model(file_path,log_filename)
        print(lrp_result)
        warmstart_time=lrp_result[4] #per depot
        nn_model_time=lrp_result[5]
        logging.info("\n\n")
        logging.info(f"Facility assignment decisions Results{lrp_result}")
        lrp_ed=datetime.now()
        flp_dict={}
        for j in range(len(lrp_result[0])):
            if lrp_result[0][j]>0.5:
                ls=[]
                for i in range(len(lrp_result[1][j])):          
                    if lrp_result[1][j][i]>0.5:
                        ls.append(i)
                
                flp_dict[j]=ls
        rout_dist={}
        fac_cust_dem={}
        cust_dem_fac={}
        for f in flp_dict:
            ls1=[]
            ls2=[]
            dem_sum=0
            for c in (flp_dict[f]):
                ls1.append(ans[3][c])
                dem_sum=dem_sum+ans[6][c]
                ls2.append(ans[6][c])
            ls1.insert(0,ans[2][f])
            ls2.insert(0,0)
            rout_dist[f]=ls1
            fac_cust_dem[f]=dem_sum
            cust_dem_fac[f]=ls2
        ass_result=[lrp_result[2],flp_dict,rout_dist,fac_cust_dem,cust_dem_fac]

        for depot_id, customers in ass_result[1].items():
            depot_coords = [ass_result[2][depot_id][0]]  # First entry is depot coordinates
            customer_coords = ass_result[2][depot_id][1:]  # Customer coordinates
            customer_demands = ass_result[4][depot_id]  # Customer demands
            vehicle_capacity = ans[4][0]  
            depot_customers = [(x[0], x[1]) for x in customer_coords]
            num_customers = len(customers)
            # filename = f"cvrp_instance_depot_{os.path.basename(file_path)}_{depot_id}.txt"
            filename = f"cvrp_instance_depot_{os.path.basename(file_path).split('.')[0]}_depot_{depot_id}_customers_{num_customers}.txt"
            output_file_path = os.path.join(DIL_instances, filename)

            write_to_txt_cvrplib_format(depot_id, depot_customers, depot_coords, customer_demands, output_file_path, vehicle_capacity)

        ve_st=datetime.now()
        print("Running vrpeasy function")
        logging.info(f"The Input to vRP easy {ass_result}")

        print(f"The Input to vRP easy {ass_result}")
        logging.info(f"ANS (Data parsed from .dat file): {ans}")

        vrpeasy_solver=dataCvrp(ans,ass_result)
        vrp_easy_results=vrpeasy_solver.runVRPeasy()
        logging.debug("VRP Results",vrp_easy_results)
        ve_ed=datetime.now()

        od=len(flp_dict)
        
        lrp_exec=((lrp_ed-lrp_st).total_seconds())/od
        warmstart_time=warmstart_time/od
        tot_ve_exec=(ve_ed-ve_st).total_seconds()
        ve_exec=((ve_ed-ve_st).total_seconds())/od
        
        # Close the log file
        logging.shutdown()

        # Create a new row
        new_row = [os.path.basename(file_path),lrp_result[2],lrp_result[3],lrp_result[2]+lrp_result[3],lrp_exec,warmstart_time,nn_model_time,vrp_easy_results[1], vrp_easy_results[0],ve_exec,tot_ve_exec,vrp_easy_results[7]]  # Replace with your data
        
        # Append the new row to the worksheet
        worksheet.append(new_row)

        # Save the modified Excel file
        workbook.save(existing_excel_file)


